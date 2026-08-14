from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATTERNS = ("Contador de renuncias*.xlsx", "contador de renuncias,*")
OUTPUT = ROOT / "data" / "cases.json"
PUBLIC_DOWNLOAD = ROOT / "data" / "base_renuncias_descarga_publica.xlsx"
SOURCE_OVERRIDES = ROOT / "data" / "source_overrides.json"
GOVERNMENT_START = date(2026, 3, 11)

INCLUDE_RECOMMENDATIONS = {
    "confirmed",
    "add_to_core_counter",
    "confirmed_from_raw_3",
    "appointment_not_effective",
}

EXCLUDE_RECOMMENDATIONS = {
    "do_not_count_pending_only",
    "pending_or_unconfirmed",
    "pending_or_paused",
}

VALID_CHILE_MINISTRIES = {
    "Ministerio del Interior",
    "Ministerio de Relaciones Exteriores",
    "Ministerio de Defensa Nacional",
    "Ministerio de Hacienda",
    "Ministerio Secretaría General de la Presidencia",
    "Ministerio Secretaría General de Gobierno",
    "Ministerio de Economía, Fomento y Turismo",
    "Ministerio de Desarrollo Social y Familia",
    "Ministerio de Educación",
    "Ministerio de Justicia y Derechos Humanos",
    "Ministerio del Trabajo y Previsión Social",
    "Ministerio de Obras Públicas",
    "Ministerio de Salud",
    "Ministerio de Vivienda y Urbanismo",
    "Ministerio de Agricultura",
    "Ministerio de Minería",
    "Ministerio de Transportes y Telecomunicaciones",
    "Ministerio de Bienes Nacionales",
    "Ministerio de Energía",
    "Ministerio del Medio Ambiente",
    "Ministerio del Deporte",
    "Ministerio de la Mujer y la Equidad de Género",
    "Ministerio de las Culturas, las Artes y el Patrimonio",
    "Ministerio de Ciencia, Tecnología, Conocimiento e Innovación",
    "Ministerio de Seguridad Pública",
}

MINISTRY_ALIASES = {
    "agricultura": "Ministerio de Agricultura",
    "bienes nacionales": "Ministerio de Bienes Nacionales",
    "ciencia": "Ministerio de Ciencia, Tecnología, Conocimiento e Innovación",
    "ciencia tecnologia conocimiento e innovacion": "Ministerio de Ciencia, Tecnología, Conocimiento e Innovación",
    "culturas": "Ministerio de las Culturas, las Artes y el Patrimonio",
    "culturas artes y patrimonio": "Ministerio de las Culturas, las Artes y el Patrimonio",
    "culturas las artes y el patrimonio": "Ministerio de las Culturas, las Artes y el Patrimonio",
    "desarrollo social y familia": "Ministerio de Desarrollo Social y Familia",
    "desarrollo social y familia mujer y equidad de genero": "Ministerio de Desarrollo Social y Familia",
    "deporte": "Ministerio del Deporte",
    "economia": "Ministerio de Economía, Fomento y Turismo",
    "economia fomento y turismo": "Ministerio de Economía, Fomento y Turismo",
    "economia y mineria": "Ministerio de Economía, Fomento y Turismo",
    "educacion": "Ministerio de Educación",
    "energia": "Ministerio de Energía",
    "hacienda": "Ministerio de Hacienda",
    "interior": "Ministerio del Interior",
    "justicia y derechos humanos": "Ministerio de Justicia y Derechos Humanos",
    "medio ambiente": "Ministerio del Medio Ambiente",
    "mineria": "Ministerio de Minería",
    "mujer y equidad de genero": "Ministerio de la Mujer y la Equidad de Género",
    "obras publicas": "Ministerio de Obras Públicas",
    "relaciones exteriores": "Ministerio de Relaciones Exteriores",
    "salud": "Ministerio de Salud",
    "secretaria general de gobierno": "Ministerio Secretaría General de Gobierno",
    "seguridad": "Ministerio de Seguridad Pública",
    "seguridad publica": "Ministerio de Seguridad Pública",
    "trabajo": "Ministerio del Trabajo y Previsión Social",
    "trabajo y prevision social": "Ministerio del Trabajo y Previsión Social",
    "transportes y telecomunicaciones": "Ministerio de Transportes y Telecomunicaciones",
    "vivienda y urbanismo": "Ministerio de Vivienda y Urbanismo",
}

NEW_COUNT_STATUSES = {"confirmed_named"}
EXCLUDED_PERSON_KEYS = {"marcelo araya"}
EXCLUDED_CASE_IDS = {"pw-090-marcelo-araya"}
SPANISH_MONTHS = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}


def strip_accents(value: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "none":
        return None
    return re.sub(r"\s+", " ", text)


def load_source_overrides() -> dict[str, dict[str, str | None]]:
    if not SOURCE_OVERRIDES.exists():
        return {}
    return json.loads(SOURCE_OVERRIDES.read_text(encoding="utf-8"))


def latest_workbook() -> Path:
    candidates = []
    for pattern in WORKBOOK_PATTERNS:
        candidates.extend(
            path
            for path in (ROOT / "data").glob(pattern)
            if path.suffix.lower() in {".xlsx", ".xlsm"} and not path.name.startswith("~$")
        )
    if not candidates:
        raise FileNotFoundError("No se encontro una base contador de renuncias en data/")
    return max(candidates, key=workbook_sort_key)


def workbook_sort_key(path: Path) -> tuple[date, float]:
    return date_from_workbook_name(path.name) or date.min, path.stat().st_mtime


def date_from_workbook_name(name: str) -> date | None:
    text = normalize_lookup_key(name)
    match = re.search(r"\b(\d{1,2})\s+(?:de\s+)?([a-z]+)\b", text)
    if not match:
        return None
    month = SPANISH_MONTHS.get(match.group(2))
    if not month:
        return None
    return date(2026, month, int(match.group(1)))


def existing_sources_by_name() -> dict[str, dict[str, str | None]]:
    if not OUTPUT.exists():
        return {}
    payload = json.loads(OUTPUT.read_text(encoding="utf-8"))
    return {
        normalize_lookup_key(case.get("person_name")): case.get("source") or {}
        for case in payload.get("cases", [])
        if case.get("person_name") and case.get("source")
    }


def iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel's Windows date system starts at 1899-12-30.
        return (date(1899, 12, 30) + timedelta(days=int(value))).isoformat()
    text = clean_text(value)
    if not text:
        return None
    parsed = datetime.fromisoformat(text.replace(" 00:00:00", ""))
    return parsed.date().isoformat()


def numeric_counter(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = clean_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def office_level(raw: str | None) -> str:
    text = strip_accents((raw or "").lower())
    if re.search(r"\bministro\b|\bministra\b", text):
        return "minister"
    if re.search(r"\bsubsecretario\b|\bsubsecretaria\b", text):
        return "subsecretary"
    if "seremi" in text:
        return "seremi"
    if "director/a nacional" in text or "directora nacional" in text or "director nacional" in text:
        return "national_director"
    if "director/a regional" in text or "directora regional" in text or "director regional" in text:
        return "regional_director"
    if "subdirector" in text:
        return "deputy_director"
    if "jefe" in text or "jefa" in text:
        return "division_head"
    return "other"


def slugify(value: str | None) -> str:
    text = strip_accents((value or "otros cargos").lower())
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text or "otros_cargos"


def normalize_lookup_key(value: str | None) -> str:
    text = strip_accents((value or "").lower())
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def ministry_from_master(row: dict[str, Any]) -> str:
    raw = clean_text(row.get("ministerio_master")) or clean_text(row.get("ministry"))
    key = normalize_lookup_key(raw)
    ministry = MINISTRY_ALIASES.get(key)
    if ministry not in VALID_CHILE_MINISTRIES:
        case_id = clean_text(row.get("master_id")) or "sin_id"
        raise ValueError(f"Ministerio master invalido en {case_id}: {raw!r}")
    return ministry


def is_excluded_case(row: dict[str, Any], name: str | None) -> bool:
    case_id = clean_text(row.get("master_id")) or clean_text(row.get("id"))
    return (case_id in EXCLUDED_CASE_IDS) or (normalize_lookup_key(name) in EXCLUDED_PERSON_KEYS)


def office_level_from_rank(rank: str | None) -> str:
    text = normalize_lookup_key(rank)
    if text in {"ministro", "ministra"}:
        return "minister"
    if text in {"subsecretario", "subsecretaria"}:
        return "subsecretary"
    if text == "seremi":
        return "seremi"
    if "director nacional" in text or "directora nacional" in text or text in {"superintendente", "superintendenta"}:
        return "national_director"
    if "director regional" in text or "directora regional" in text:
        return "regional_director"
    if "subdirector" in text:
        return "deputy_director"
    if "jefatura" in text:
        return "division_head"
    return "other"


def cargo_group_from_level(level: str) -> str:
    return {
        "minister": "Ministro/a",
        "subsecretary": "Subsecretario/a",
        "seremi": "Seremi",
        "national_director": "Director/a nacional",
        "regional_director": "Director/a regional",
        "deputy_director": "Subdirector/a",
        "division_head": "Jefatura",
    }.get(level, "Otros cargos")


def region_group(value: str) -> str:
    normalized = strip_accents(value.lower())
    if normalized == "nacional":
        return "Gobierno central"
    if normalized in {"araucania", "la araucania"}:
        return "La Araucanía"
    if normalized in {"metropolitana", "region metropolitana"}:
        return "Región Metropolitana"
    return value


def exit_type(raw: str | None) -> str:
    text = strip_accents((raw or "").lower())
    if "movimiento" in text:
        return "internal_movement"
    if "nombramiento" in text and ("sin_efecto" in text or "sin efecto" in text):
        return "appointment_not_effective"
    if "solicitada" in text or "no_voluntaria" in text or "no voluntaria" in text:
        return "resignation_requested"
    if "remocion" in text or "desvinculacion" in text:
        return "removed"
    if "voluntaria" in text:
        return "voluntary_resignation"
    if "renuncia" in text or "aceptada" in text:
        return "resignation"
    return "unknown"


def updated_at_from_summary(wb: Any, fallback: date) -> str:
    if "Resumen" not in wb.sheetnames:
        return fallback.isoformat()
    ws = wb["Resumen"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if clean_text(row[0]) == "Actualizado" and row[1]:
            return iso_date(row[1]) or fallback.isoformat()
    return fallback.isoformat()


def government_weeks_elapsed(start: date, end: date) -> int:
    if end < start:
        return 0
    return ((end - start).days // 7) + 1


def reason_category(row: dict[str, Any], summary: str | None) -> str:
    combined = strip_accents(
        " ".join(
            clean_text(v) or ""
            for v in [
                row.get("tipo_salida_factiva"),
                row.get("razon_renunciaskast"),
                row.get("resumen_factiva"),
                row.get("titular_factiva"),
                summary,
            ]
        ).lower()
    )
    if any(term in combined for term in ["denuncia", "investigacion penal", "acoso", "connotacion sexual"]):
        return "judicial_or_formal_complaint"
    if any(term in combined for term in ["test de drogas", "examen de drogas"]):
        return "drug_test_or_compliance"
    if any(term in combined for term in ["nombramiento", "requisito", "titulo", "semestres", "sin efecto"]):
        return "requirements_or_appointment"
    if any(term in combined for term in ["falta de gestion", "cuestionad", "polemica", "posteos"]):
        return "public_management_questioning"
    if any(term in combined for term in ["tension", "choque", "conflicto"]):
        return "internal_conflict"
    if "motivos personales" in combined or "razones personales" in combined:
        return "personal_reasons"
    return "not_specified"


def source_for(row: dict[str, Any], overrides: dict[str, dict[str, str | None]]) -> dict[str, str | None]:
    override = overrides.get(clean_text(row.get("master_id")) or "")
    url = clean_text(row.get("url_renunciaskast"))
    outlet = clean_text(row.get("medio_renunciaskast")) or clean_text(row.get("medio_factiva"))
    title = clean_text(row.get("titular_factiva"))
    if override and not url:
        url = clean_text(override.get("url"))
        outlet = clean_text(override.get("outlet")) or outlet
        title = clean_text(override.get("title")) or title
    return {"outlet": outlet or "Fuente", "url": url, "title": title}


def source_for_new(row: dict[str, Any], preserved: dict[str, dict[str, str | None]], rk_sources: dict[str, dict[str, str | None]]) -> dict[str, str | None]:
    name_key = normalize_lookup_key(clean_text(row.get("name")))
    if name_key in rk_sources:
        return rk_sources[name_key]
    if name_key in preserved:
        return preserved[name_key]
    article = clean_text(row.get("source_article")) or clean_text(row.get("source_article_abs"))
    title = clean_text(row.get("source_article"))
    return {"outlet": "Fuente", "url": None, "title": title or article}


def renunciaskast_sources(wb: Any) -> dict[str, dict[str, str | None]]:
    if "RenunciasKast" not in wb.sheetnames:
        return {}
    ws = wb["RenunciasKast"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    index = {h: i for i, h in enumerate(headers) if h}
    sources = {}
    for raw_row in rows[1:]:
        row = {h: raw_row[i] if i < len(raw_row) else None for h, i in index.items()}
        name = clean_text(row.get("n"))
        if not name:
            continue
        sources[normalize_lookup_key(name)] = {
            "outlet": clean_text(row.get("medio")) or "Fuente",
            "url": clean_text(row.get("url")),
            "title": clean_text(row.get("razon")) or clean_text(row.get("medio")),
        }
    return sources


def build() -> dict[str, Any]:
    workbook = latest_workbook()
    wb = load_workbook(workbook, read_only=True, data_only=True, keep_vba=workbook.suffix.lower() == ".xlsm")
    source_overrides = load_source_overrides()
    preserved_sources = existing_sources_by_name()
    rk_sources = renunciaskast_sources(wb)
    ws = wb["Master"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    index = {h: i for i, h in enumerate(headers) if h}
    cases: list[dict[str, Any]] = []

    is_new_schema = {"id", "name", "position", "count_status"}.issubset(set(headers))

    for raw_row in rows[1:]:
        row = {h: raw_row[i] if i < len(raw_row) else None for h, i in index.items()}
        if is_new_schema:
            if clean_text(row.get("count_status")) not in NEW_COUNT_STATUSES:
                continue
            name = clean_text(row.get("name"))
            if not name:
                continue
            if is_excluded_case(row, name):
                continue
            date_value = iso_date(row.get("exit_date"))
            summary = clean_text(row.get("evidence")) or clean_text(row.get("notes")) or "Sin motivo publico detallado."
            raw_cargo = clean_text(row.get("position")) or clean_text(row.get("rank"))
            level = office_level_from_rank(clean_text(row.get("rank")) or raw_cargo)
            cargo_group = cargo_group_from_level(level)
            region = clean_text(row.get("region")) or "Nacional"
            ministry_master = ministry_from_master(row)
            case = {
                "case_id": clean_text(row.get("id")),
                "person_name": name,
                "office_level": level,
                "cargo_group": cargo_group,
                "cargo_group_key": slugify(cargo_group),
                "office_title": raw_cargo,
                "ministerio_master": ministry_master,
                "ministry": ministry_master,
                "territory_type": "national" if region == "Nacional" else "regional",
                "region": region,
                "region_group": region_group(region),
                "exit_date": date_value,
                "exit_type": exit_type(clean_text(row.get("exit_type"))),
                "reason_category": reason_category(row, summary),
                "reason_summary": summary,
                "has_judicial_or_formal_complaint": reason_category(row, summary) == "judicial_or_formal_complaint",
                "verification_status": "verified",
                "count_recommendation": clean_text(row.get("count_status")),
                "seremi_counter": None,
                "source": source_for_new(row, preserved_sources, rk_sources),
            }
            cases.append(case)
            continue

        recommendation = clean_text(row.get("recomendacion_conteo"))
        seremi_counter = numeric_counter(row.get("Seremis Contador"))
        counted_by_master = seremi_counter is not None and recommendation not in EXCLUDE_RECOMMENDATIONS
        if recommendation not in INCLUDE_RECOMMENDATIONS and not counted_by_master:
            continue
        name = clean_text(row.get("nombre_master"))
        if not name:
            continue
        if is_excluded_case(row, name):
            continue

        date_value = iso_date(row.get("fecha_salida_master") or row.get("fecha_factiva") or row.get("fecha_renunciaskast"))
        summary = clean_text(row.get("razon_renunciaskast")) or clean_text(row.get("resumen_factiva")) or clean_text(row.get("titular_factiva"))
        raw_cargo = clean_text(row.get("cargo_master")) or clean_text(row.get("cargo"))
        cargo_group = clean_text(row.get("cargo")) or raw_cargo or "Otros cargos"
        region = clean_text(row.get("region_master")) or "Nacional"
        ministry_master = ministry_from_master(row)
        case = {
            "case_id": clean_text(row.get("master_id")),
            "person_name": name,
            "office_level": "seremi" if seremi_counter is not None else office_level(cargo_group),
            "cargo_group": cargo_group,
            "cargo_group_key": slugify(cargo_group),
            "office_title": raw_cargo,
            "ministerio_master": ministry_master,
            "ministry": ministry_master,
            "territory_type": "national" if region == "Nacional" else "regional",
            "region": region,
            "region_group": region_group(region),
            "exit_date": date_value,
            "exit_type": exit_type(clean_text(row.get("tipo_salida_factiva"))),
            "reason_category": reason_category(row, summary),
            "reason_summary": summary or "Sin motivo publico detallado.",
            "has_judicial_or_formal_complaint": reason_category(row, summary) == "judicial_or_formal_complaint",
            "verification_status": "verified" if clean_text(row.get("chequeo_manual")) == "☑" else "needs_review",
            "count_recommendation": recommendation,
            "seremi_counter": seremi_counter,
            "source": source_for(row, source_overrides),
        }
        cases.append(case)

    cases.sort(key=lambda item: item.get("exit_date") or "1900-01-01", reverse=True)
    office_counts = Counter(item["office_level"] for item in cases)
    cargo_counts = Counter(item["cargo_group"] for item in cases)
    ministry_counts = Counter(item["ministerio_master"] for item in cases)
    region_counts = Counter(item["region_group"] for item in cases)
    fallback_updated_at = date_from_workbook_name(workbook.name) or max(
        (datetime.fromisoformat(item["exit_date"]).date() for item in cases if item.get("exit_date")),
        default=date.today(),
    )
    updated_at = updated_at_from_summary(wb, fallback_updated_at)
    updated_date = datetime.fromisoformat(updated_at).date()
    weeks_elapsed = government_weeks_elapsed(GOVERNMENT_START, updated_date)
    resignations_per_week = round(len(cases) / weeks_elapsed, 3) if weeks_elapsed else None
    return {
        "metadata": {
            "title": "Renuncias Tracker",
            "updated_at": updated_at,
            "source_note": f"Base generada desde la pestaña Master de {workbook.name}.",
            "case_count": len(cases),
            "seremi_count": max((item["seremi_counter"] or 0 for item in cases if item["office_level"] == "seremi"), default=0)
            or office_counts.get("seremi", 0),
            "office_counts": dict(office_counts),
            "cargo_counts": dict(cargo_counts),
            "ministry_counts": dict(ministry_counts),
            "region_counts": dict(region_counts),
            "government_start": GOVERNMENT_START.isoformat(),
            "government_weeks_elapsed": weeks_elapsed,
            "week_basis": "Semanas corridas de gobierno desde el 11 de marzo de 2026, incluyendo la semana parcial en curso.",
            "resignations_per_week": resignations_per_week,
        },
        "cases": cases,
    }


def write_public_download(payload: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Base limpia"
    headers = ["Nombre", "cargo", "ministerio", "region", "fecha de salida", "url"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="071B3D")

    for case in payload["cases"]:
        ws.append(
            [
                case.get("person_name"),
                case.get("office_title"),
                case.get("ministerio_master"),
                case.get("region"),
                case.get("exit_date"),
                (case.get("source") or {}).get("url"),
            ]
        )
        url_cell = ws.cell(ws.max_row, 6)
        if url_cell.value:
            url_cell.hyperlink = url_cell.value
            url_cell.style = "Hyperlink"

    widths = [28, 44, 42, 24, 16, 68]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(PUBLIC_DOWNLOAD)


if __name__ == "__main__":
    payload = build()
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT} with {payload['metadata']['case_count']} cases")
    write_public_download(payload)
    print(f"Wrote {PUBLIC_DOWNLOAD}")
